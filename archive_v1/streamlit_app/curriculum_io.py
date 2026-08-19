"""
The one way curriculum gets into Flokus Academy, and the one way it comes out.

WHY THERE IS ONLY ONE
---------------------
v1 had three: a Quick Add form, per-day plus-buttons hidden in the weekly grid,
and a "Master Curriculum Scheduler" that regenerated the entire year from
hardcoded tables in curriculum_data.py. The third was the dangerous one -- one
click re-scheduled 677 assignments on top of whatever was already there, and
because it read from Python source rather than a file, the actual plan could
only be changed by editing code.

Now: a spreadsheet goes in, a spreadsheet comes out, and single one-off lessons
use the Quick Add form. Nothing else writes curriculum.

THE FORMAT
----------
The layout already sitting in research_and_development/
Flokus_Curriculum_Template.xlsx -- one sheet per program, one row per lesson.
This is the same format V2 ingests, so a file that imports here imports there.
That is the whole point of matching it rather than inventing a v1-only format.

v1 stores a subset of those columns (its tasks table has no unit, priority or
UFA fields). The rest pass through untouched so nothing is lost on a round
trip -- v1 simply doesn't act on them yet.

LESSONS AND ROUTINES ARE DIFFERENT ROWS
---------------------------------------
A lesson happens once: "TT Vol 1 Ch 3", "BA 3A Ch 2 - Guide, part 1". A routine
is a weekly slot that repeats all year under the same name: "Chess - Tactics &
Play" every Wednesday, 36 times.

Writing a routine out 36 times would mean 36 rows whose program, unit and title
are identical -- and source_key is built from exactly those three fields, so
every copy after the first would look like a duplicate of the first. It is not
a naming problem that can be patched; the two things are shaped differently and
belong in different sheets. That is why the shipped template has a Routines
sheet, and why this module reads and writes one.

IMPORTING NEVER GUESSES A DATE
------------------------------
The file says WHAT to teach and in what order. The caller says WHEN -- a start
date and the weekdays the program occupies -- and scheduling happens against
school_year.py, so an import cannot land on Thanksgiving. A date column in the
file would let the file and the calendar disagree, and the file would win.

RE-IMPORTING IS SAFE
--------------------
Every row carries a source_key. Re-importing a corrected file updates the
matching rows instead of adding a second copy of the year, and completed work
is never touched. That is what makes it safe to fix a typo in Excel and load it
again.
"""

import csv
import io
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from datetime import date, timedelta

import school_year

DB = "flokus.db"

LESSON_COLUMNS = [
    "program", "unit", "unit_status", "unit_week_start", "unit_week_end",
    "title", "description", "task_type", "priority", "sequence_order",
    "estimated_minutes", "xp_reward", "is_boss_fight", "medium",
    "dependency_mode", "day_of_week_hint", "resource_url", "workbook_pages",
    "ufa_eligible", "ufa_hours_credit", "source_key",
]

ROUTINE_COLUMNS = [
    "program", "status", "title", "cadence", "day_of_week", "starts_week",
    "ends_week", "task_type", "estimated_minutes", "xp_reward", "medium",
    "dependency_mode", "ufa_eligible", "ufa_hours_credit",
]

REQUIRED_LESSON = ("program", "title")
REQUIRED_ROUTINE = ("program", "title", "day_of_week")

VALID_TASK_TYPES = {"reading", "practice", "lesson", "project", "review",
                    "build", "assessment", ""}
VALID_PRIORITY = {"core", "standard", "optional", ""}
VALID_MEDIUM = {"offline", "online", ""}

# A sheet says "Beast Academy"; v1's tasks table says "Math (Beast Academy)".
# The mapping is explicit rather than fuzzy-matched because a near-miss would
# silently create a subject that shows up nowhere in the UI.
PROGRAM_TO_CATEGORY = {
    "beast academy": "Math (Beast Academy)",
    "brave writer": "Language Arts (Brave Writer)",
    "crunchlabs": "Science (CrunchLabs)",
    "outschool": "Science (Outschool)",
    "tuttle twins": "Social Studies (Tuttle Twins)",
    "free market rules": "Economics (Free Market Rules)",
    "synthesis": "Logic (Synthesis)",
    "chess com": "Logic (Chess.com)",
    "critical thinking co": "Logic (Critical Thinking Co.)",
}
CATEGORY_TO_PROGRAM = {
    "Math (Beast Academy)": "Beast Academy",
    "Language Arts (Brave Writer)": "Brave Writer",
    "Science (CrunchLabs)": "CrunchLabs",
    "Science (Outschool)": "Outschool",
    "Social Studies (Tuttle Twins)": "Tuttle Twins",
    "Economics (Free Market Rules)": "Free Market Rules",
    "Logic (Chess.com)": "Chess.com",
    "Logic (Critical Thinking Co.)": "Critical Thinking Co.",
    "Logic (Synthesis)": "Synthesis",
}


def slug(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def _program_token(program):
    return slug(program).replace("-", " ")


def resolve_category(program):
    return PROGRAM_TO_CATEGORY.get(_program_token(program))


def make_source_key(program, unit, title):
    return f"{slug(program)}|{slug(unit)}|{slug(title)}"


def routine_source_key(program, title, day):
    """Routines repeat, so the date is part of the identity. Without it every
    occurrence of 'Chess - Tactics & Play' would collide with every other."""
    return f"routine|{slug(program)}|{slug(title)}|{day.isoformat()}"


# ---------------------------------------------------------------------------
# Reading a file
# ---------------------------------------------------------------------------

def _norm_header(cells):
    return [re.sub(r"\s+", "_", str(c or "").strip().lower()) for c in cells]


def read_rows(uploaded_name, data):
    """Parse an .xlsx or .csv into (lessons, routines, sheet_notes).

    The header line is located rather than assumed, because the shipped
    template puts a human-readable note above it. Sheets with no recognisable
    header (READ ME, Valid values) are skipped by shape, not by name, so a
    renamed doc sheet doesn't become forty bogus errors.
    """
    lessons, routines, notes = [], [], []
    name = (uploaded_name or "").lower()

    if name.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        sheets = [("csv", [r for r in csv.reader(io.StringIO(text))])]
    elif name.endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
        except ImportError:
            return [], [], ["openpyxl is not installed -- run: pip install openpyxl"]
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            return [], [], [f"Could not open the workbook: {exc}"]
        sheets = [(sn, [list(r) for r in wb[sn].iter_rows(values_only=True)])
                  for sn in wb.sheetnames]
    else:
        return [], [], ["Unsupported file type. Upload a .xlsx or .csv."]

    for sheet_name, grid in sheets:
        header_idx = header = kind = None
        for i, raw in enumerate(grid[:8]):
            cells = _norm_header(raw)
            if "title" not in cells or "program" not in cells:
                continue
            header_idx, header = i, cells
            # A routine sheet is identified by carrying a cadence or a
            # day_of_week column -- that is what makes it a repeating slot.
            kind = "routine" if ("cadence" in cells or "day_of_week" in cells) \
                else "lesson"
            break
        if header_idx is None:
            continue

        allowed = ROUTINE_COLUMNS if kind == "routine" else LESSON_COLUMNS
        unknown = [c for c in header if c and c not in allowed]
        if unknown:
            notes.append(f"[{sheet_name}] ignoring unrecognised column(s): "
                         f"{', '.join(unknown)}")

        for offset, raw in enumerate(grid[header_idx + 1:], start=header_idx + 2):
            values = {}
            for key, cell in zip(header, raw):
                if key in allowed:
                    values[key] = "" if cell is None else str(cell).strip()
            if not any(values.get(c) for c in ("program", "title")):
                continue                                   # blank spacer row
            values["_sheet"] = sheet_name
            values["_row"] = offset
            (routines if kind == "routine" else lessons).append(values)

    if not lessons and not routines and not notes:
        notes.append("No curriculum rows found. Every sheet needs a header line "
                     "containing at least 'program' and 'title'.")
    return lessons, routines, notes


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _check_numbers(r, where, cols, errors):
    for col in cols:
        v = r.get(col)
        if v:
            try:
                if int(float(v)) < 0:
                    errors.append(f"{where} {col} cannot be negative")
            except ValueError:
                errors.append(f"{where} {col} must be a whole number, got {v!r}")


def validate(lessons, routines):
    """Return (lessons, routines, errors, warnings). Errors are row-numbered.

    Everything checkable is checked before anything is written, so a bad file
    fails wholly and visibly rather than half-importing.
    """
    errors, warnings = [], []
    known = ", ".join(sorted(CATEGORY_TO_PROGRAM.values()))
    seen = {}

    for r in lessons:
        where = f"[{r['_sheet']} row {r['_row']}]"
        for col in REQUIRED_LESSON:
            if not r.get(col):
                errors.append(f"{where} missing required column '{col}'")
        program = r.get("program", "")
        if program and not resolve_category(program):
            errors.append(f"{where} unknown program {program!r}. Known: {known}")

        tt = (r.get("task_type") or "").lower()
        if tt and tt not in VALID_TASK_TYPES:
            warnings.append(f"{where} unusual task_type {tt!r} -- kept as-is")
        pr = (r.get("priority") or "").lower()
        if pr and pr not in VALID_PRIORITY:
            errors.append(f"{where} priority must be core/standard/optional, got {pr!r}")
        md = (r.get("medium") or "").lower()
        if md and md not in VALID_MEDIUM:
            errors.append(f"{where} medium must be offline/online, got {md!r}")
        _check_numbers(r, where, ("xp_reward", "estimated_minutes",
                                  "sequence_order"), errors)

        key = r.get("source_key") or make_source_key(
            program, r.get("unit", ""), r.get("title", ""))
        r["source_key"] = key
        if key in seen:
            errors.append(
                f"{where} duplicates {seen[key]} -- same program, unit and "
                f"title. If this is a lesson, give it a distinct title (e.g. "
                f"'... - Part 2'). If it is a weekly habit, it belongs on a "
                f"Routines sheet instead, listed once.")
        else:
            seen[key] = where

    for r in routines:
        where = f"[{r['_sheet']} row {r['_row']}]"
        for col in REQUIRED_ROUTINE:
            if not r.get(col):
                errors.append(f"{where} routine missing required column '{col}'")
        program = r.get("program", "")
        if program and not resolve_category(program):
            errors.append(f"{where} unknown program {program!r}. Known: {known}")
        dow = (r.get("day_of_week") or "").strip().title()[:3]
        if dow and dow not in school_year.WEEKDAY_NAMES:
            errors.append(f"{where} day_of_week must be Mon-Fri, got "
                          f"{r.get('day_of_week')!r}")
        else:
            r["day_of_week"] = dow
        _check_numbers(r, where, ("xp_reward", "estimated_minutes",
                                  "starts_week", "ends_week"), errors)
        if (r.get("status") or "active").lower() not in ("active", "paused", ""):
            warnings.append(f"{where} status {r.get('status')!r} -- treating as active")

    return lessons, routines, errors, warnings


def to_task(r):
    """Map a validated lesson row onto the columns v1's tasks table has."""
    def as_int(v, default):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default
    return {
        "title": r["title"],
        "category": resolve_category(r.get("program", "")),
        "xp_reward": as_int(r.get("xp_reward"), 10),
        "medium": "Online" if (r.get("medium") or "").lower() == "online" else "Offline",
        "video_url": r.get("resource_url", "") or "",
        "source_key": r["source_key"],
    }


# ---------------------------------------------------------------------------
# Scheduling and committing
# ---------------------------------------------------------------------------

def plan_dates(count, weekdays, start):
    """Which school days `count` lessons will land on. Never a day off."""
    return school_year.slots(weekdays, start=start, limit=count)


def expand_routine(r):
    """Turn one routine definition into its occurrences across the year."""
    _, week_starts = school_year.build_calendar()
    def as_int(v, default):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default
    first = max(1, as_int(r.get("starts_week"), 1))
    last = min(len(week_starts), as_int(r.get("ends_week"), len(week_starts)))
    wd = school_year.WEEKDAY_NAMES[r["day_of_week"]]
    days = []
    for idx in range(first - 1, last):
        day = week_starts[idx] + timedelta(days=wd)
        if school_year.day_off_reason(day) is None:
            days.append(day)
    return days


def preview(lessons, routines, weekdays, start, replace_program=True, db=DB):
    """What a commit would do, without doing it.

    Includes the projected daily load, because the failure mode here is quiet
    and expensive: choosing Mon-Thu for a program that actually runs Tue+Thu
    schedules it correctly by every other measure and simply makes Sonny's days
    too long. Better to see "14 days would go over 5 assignments" before
    committing than to find out in November.
    """
    dates = plan_dates(len(lessons), weekdays, start)
    routine_days = {r["title"]: expand_routine(r) for r in routines
                    if r.get("day_of_week")}
    categories = {resolve_category(r.get("program", ""))
                  for r in lessons + routines if resolve_category(r.get("program", ""))}

    load = defaultdict(int)
    conn = sqlite3.connect(db)
    try:
        for day_str, n in conn.execute(
                "SELECT task_date, COUNT(*) FROM tasks GROUP BY task_date"):
            load[date.fromisoformat(day_str)] = n
        if replace_program:
            for cat in categories:
                for day_str, n in conn.execute(
                        "SELECT task_date, COUNT(*) FROM tasks WHERE category = ? "
                        "AND is_completed = 0 AND task_date >= ? GROUP BY task_date",
                        (cat, start.isoformat())):
                    load[date.fromisoformat(day_str)] -= n
    finally:
        conn.close()

    for day in dates[:len(lessons)]:
        load[day] += 1
    over = sorted(d for d, n in load.items() if n > school_year.MAX_TASKS_PER_DAY)

    return {
        "lesson_count": len(lessons),
        "lesson_slots": len(dates),
        "lesson_first": dates[0] if dates else None,
        "lesson_last": dates[len(lessons) - 1] if len(dates) >= len(lessons) and lessons else None,
        "short_by": max(0, len(lessons) - len(dates)),
        "routines": {t: len(d) for t, d in routine_days.items()},
        "programs": sorted(categories),
        "over_cap_days": len(over),
        "over_cap_first": over[0] if over else None,
        "busiest": max(load.values()) if load else 0,
    }


def commit(lessons, routines, weekdays, start, replace_program=True, db=DB):
    """Write an import atomically. Returns a summary dict.

    Atomic on purpose: a partially-applied curriculum is worse than a rejected
    one, because it looks like it worked.
    """
    tasks = [to_task(r) for r in lessons]
    dates = plan_dates(len(tasks), weekdays, start)
    if len(dates) < len(tasks):
        raise ValueError(
            f"Only {len(dates)} school days left on the chosen weekdays from "
            f"{start:%b %d, %Y}, but the file has {len(tasks)} lessons. Pick an "
            f"earlier start date, or add a weekday.")

    routine_tasks = []
    for r in routines:
        if (r.get("status") or "active").lower() == "paused":
            continue
        cat = resolve_category(r.get("program", ""))
        try:
            xp = int(float(r.get("xp_reward") or 10))
        except ValueError:
            xp = 10
        for day in expand_routine(r):
            if day < start:
                continue
            routine_tasks.append(({
                "title": r["title"],
                "category": cat,
                "xp_reward": xp,
                "medium": "Online" if (r.get("medium") or "").lower() == "online" else "Offline",
                "video_url": "",
                "source_key": routine_source_key(r["program"], r["title"], day),
            }, day))

    categories = {t["category"] for t in tasks} | {t["category"] for t, _ in routine_tasks}
    conn = sqlite3.connect(db)
    try:
        cur = conn.cursor()
        _ensure_source_key_column(cur)
        removed = 0
        if replace_program:
            for cat in categories:
                # Only unfinished rows are cleared. Anything Sonny has already
                # completed stays on the record -- a re-import changes the plan,
                # it does not rewrite what he actually did.
                removed += cur.execute(
                    "DELETE FROM tasks WHERE category = ? AND is_completed = 0 "
                    "AND task_date >= ?", (cat, start.isoformat())).rowcount

        # A routine yields its slot when the day is already full. A CrunchLabs
        # build day IS that week's applied-STEM session and a Book Party IS that
        # week's Brave Writer Friday, so re-adding the habit alongside the event
        # pushes those days to six -- it turns the best days of the year into
        # the heaviest ones. Lessons are never skipped this way: they carry a
        # sequence that has to be delivered, so an overfull day is reported
        # instead of silently thinned.
        load = defaultdict(int)
        for day_str, n in cur.execute(
                "SELECT task_date, COUNT(*) FROM tasks GROUP BY task_date"):
            load[date.fromisoformat(day_str)] = n
        for _, day in list(zip(tasks, dates)):
            load[day] += 1

        kept_routines, skipped_full = [], 0
        for task, day in routine_tasks:
            if load[day] >= school_year.MAX_TASKS_PER_DAY:
                already = cur.execute(
                    "SELECT 1 FROM tasks WHERE source_key = ?",
                    (task["source_key"],)).fetchone()
                if not already:
                    skipped_full += 1
                    continue
            load[day] += 1
            kept_routines.append((task, day))
        routine_tasks = kept_routines

        updated = inserted = 0
        for task, day in list(zip(tasks, dates)) + routine_tasks:
            existing = cur.execute(
                "SELECT id FROM tasks WHERE source_key = ? AND source_key != '' "
                "AND is_completed = 0", (task["source_key"],)).fetchone()
            if existing:
                cur.execute(
                    "UPDATE tasks SET title=?, category=?, task_date=?, "
                    "video_url=?, xp_reward=?, medium=? WHERE id=?",
                    (task["title"], task["category"], day.isoformat(),
                     task["video_url"], task["xp_reward"], task["medium"],
                     existing[0]))
                updated += 1
            else:
                cur.execute(
                    "INSERT INTO tasks (title, category, task_date, video_url, "
                    "xp_reward, is_completed, is_boss_fight, task_summary, "
                    "focus_minutes, actual_completion_date, medium, source_key) "
                    "VALUES (?,?,?,?,?,0,0,'',0,'',?,?)",
                    (task["title"], task["category"], day.isoformat(),
                     task["video_url"], task["xp_reward"], task["medium"],
                     task["source_key"]))
                inserted += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    all_days = dates[:len(tasks)] + [d for _, d in routine_tasks]
    return {"inserted": inserted, "updated": updated, "removed": removed,
            "skipped_full": skipped_full, "programs": sorted(categories),
            "first": min(all_days) if all_days else None,
            "last": max(all_days) if all_days else None}


def _ensure_source_key_column(cur):
    try:
        cur.execute("SELECT source_key FROM tasks LIMIT 1")
    except sqlite3.OperationalError:
        cur.execute("ALTER TABLE tasks ADD COLUMN source_key TEXT DEFAULT ''")


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def split_lessons_and_routines(rows):
    """Decide which scheduled rows describe a repeating slot.

    A routine is a title that recurs under one program and always on the same
    weekday -- that is precisely what "weekly habit" means in this data. Every
    other title happens once and is a lesson. Detecting it from the shape of
    the data rather than from a list of known names means a routine added next
    year is classified correctly without anyone updating this file.
    """
    counts = Counter((cat, title) for cat, title, *_ in rows)
    weekdays = defaultdict(set)
    for cat, title, day, *_ in rows:
        weekdays[(cat, title)].add(date.fromisoformat(day).weekday())
    is_routine = {k for k, n in counts.items()
                  if n > 1 and len(weekdays[k]) == 1}
    return ([r for r in rows if (r[0], r[1]) not in is_routine],
            [r for r in rows if (r[0], r[1]) in is_routine],
            weekdays)


def export_workbook(path, db=DB):
    """Write the live schedule back out in the import format. The result
    re-imports cleanly, which is the only useful definition of 'exported
    correctly'."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    conn = sqlite3.connect(db)
    _ensure_source_key_column(conn.cursor())
    rows = list(conn.execute(
        "SELECT category, title, task_date, xp_reward, medium, video_url, "
        "COALESCE(source_key,'') FROM tasks ORDER BY category, task_date, id"))
    conn.close()

    lesson_rows, routine_rows, weekdays = split_lessons_and_routines(rows)
    _, week_starts = school_year.build_calendar()
    day_names = {v: k for k, v in school_year.WEEKDAY_NAMES.items()}

    by_program = defaultdict(list)
    for r in lesson_rows:
        by_program[CATEGORY_TO_PROGRAM.get(r[0], r[0])].append(r)

    routines = {}
    for cat, title, day, xp, medium, url, skey in routine_rows:
        entry = routines.setdefault((cat, title), {
            "program": CATEGORY_TO_PROGRAM.get(cat, cat), "title": title,
            "xp": xp, "medium": medium, "days": []})
        entry["days"].append(date.fromisoformat(day))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    s = school_year.summary()
    info = wb.create_sheet("READ ME")
    for line in [
        ("Flokus Academy - 2026-27 curriculum, exported from flokus.db", ""),
        ("", ""),
        ("School year", f"{s['start']:%a %d %b %Y} - {s['end']:%a %d %b %Y}"),
        ("Instructional weeks", s["weeks"]),
        ("School days", s["school_days"]),
        ("Days off", s["days_off"]),
        ("Assignments", len(rows)),
        ("  of which lessons", len(lesson_rows)),
        ("  of which routine occurrences", len(routine_rows)),
        ("Routine definitions", len(routines)),
        ("", ""),
        ("Format", "One sheet per program, one row per lesson, teaching order."),
        ("", "Columns match Flokus_Curriculum_Template.xlsx and the V2 import"),
        ("", "spec, so this file loads into either."),
        ("", ""),
        ("Lessons vs routines", "A lesson happens once. A routine is a weekly"),
        ("", "slot repeating under the same name all year, so it is listed"),
        ("", "ONCE on the Routines sheet with its weekday - not 36 times."),
        ("", ""),
        ("No date column - by design", "The file says WHAT and in what order."),
        ("", "The importer asks WHEN (start date + weekdays) and schedules"),
        ("", "against the school calendar, so an import cannot land on a"),
        ("", "holiday. A date column could disagree with the calendar."),
        ("", ""),
        ("Re-importing is safe", "source_key = program|unit|title. Re-importing"),
        ("", "a corrected file updates those rows instead of duplicating them."),
        ("", "Completed work is never touched by an import."),
    ]:
        info.append(list(line))
    info.column_dimensions["A"].width = 30
    info.column_dimensions["B"].width = 78
    info["A1"].font = Font(bold=True, size=13)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")

    def style(ws):
        for cell in ws[1]:
            cell.font, cell.fill = head_font, head_fill
        ws.freeze_panes = "A2"

    for program, items in sorted(by_program.items()):
        ws = wb.create_sheet(program[:31])
        ws.append(LESSON_COLUMNS)
        style(ws)
        for seq, (cat, title, day, xp, medium, url, skey) in enumerate(items, 1):
            ws.append([program, "", "active", "", "", title, "", "", "", seq,
                       "", xp, "no", (medium or "Offline").lower(), "", "",
                       url or "", "", "yes", "",
                       skey or make_source_key(program, "", title)])
        for col, width in (("A", 24), ("F", 64), ("U", 46)):
            ws.column_dimensions[col].width = width

    if routines:
        ws = wb.create_sheet("Routines")
        ws.append(ROUTINE_COLUMNS)
        style(ws)
        for (cat, title), e in sorted(routines.items()):
            days = sorted(e["days"])
            wd = day_names[days[0].weekday()]
            first_week = min(i for i, m in enumerate(week_starts, 1)
                             if m <= days[0] <= m + timedelta(days=4))
            last_week = max(i for i, m in enumerate(week_starts, 1)
                            if m <= days[-1] <= m + timedelta(days=4))
            ws.append([
                e["program"], "active", title, "weekly", wd,
                first_week if first_week != 1 else "",
                last_week if last_week != len(week_starts) else "",
                "", "", e["xp"], (e["medium"] or "Offline").lower(), "", "yes", "",
            ])
        for col, width in (("A", 22), ("C", 52)):
            ws.column_dimensions[col].width = width

    wb.save(path)
    return {"programs": sorted(by_program), "lessons": len(lesson_rows),
            "routines": len(routines), "rows": len(rows), "path": path}


if __name__ == "__main__":
    out = export_workbook("Flokus_Curriculum_2026-27_EXPORT.xlsx")
    print(f"{out['rows']} rows -> {out['path']}")
    print(f"  {out['lessons']} lessons across {len(out['programs'])} programs")
    print(f"  {out['routines']} routine definitions")
    for p in out["programs"]:
        print("   ", p)
